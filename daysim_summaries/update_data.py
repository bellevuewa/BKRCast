# import libraries
import os
import csv
import datetime
import pandas as pd
import numpy as np
import xlrd
import openpyxl
from win32com.client import DispatchEx
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.styles import Color

#data folder
wd = r"D:/SQFTtoJobsTest/BKRCast_V1-3LU_Test19/daysim_summaries"

infile_names = ["DayPattern.xlsm"]
infile_names = ["DayPattern.xlsm",
                "WrkLocation.xlsm", "SchLocation.xlsm",
                #"VehAvailability.xlsm",
                "TourDestination_Escort.xlsm","TourDestination_Meal.xlsm","TourDestination_PerBus.xlsm",
                "TourDestination_Shop.xlsm","TourDestination_SocRec.xlsm","TourDestination_WrkBased.xlsm",
                "TripDestination.xlsm",
                "TourMode.xlsm","TripMode.xlsm"]
sheet_names = ["Summary",
               "TLFD","TLFD",
               #"calibration",
               "TLFD","TLFD","TLFD",
               "TLFD","TLFD","TLFD",
               "TLFD",
               "AllPurposes","All Purposes"]
outfile_names = ["DayPattern_bkrcast.xlsm",
                 "WrkLocation_bkrcast.xlsm", "SchLocation_bkrcast.xlsm",
                 #"VehAvailability.xlsx",
                "TourDestination_Escort_bkrcast.xlsm","TourDestination_Meal_bkrcast.xlsm","TourDestination_PerBus_bkrcast.xlsm",
                "TourDestination_Shop_bkrcast.xlsm","TourDestination_SocRec_bkrcast.xlsm","TourDestination_WrkBased_bkrcast.xlsm",
                "TripDestination_bkrcast.xlsm",
                "TourMode_bkrcast.xlsm","TripMode_bkrcast.xlsm"]
purpose = ["Home","Work", "School", "Escort", "PerBus", "Shop", "Meal", "SocRec"]
run_names = ["bkrcast_all","bkrcast_inbkr","bkrcast_outbkr"]
#run_names = ["bkrcast_all","bkrcast_inbkr","bkrcast_outbkr","soundcast"]

def add_chart(sheet, chart_name, x_axis_name, y_axis_name, start_row, end_row, start_col, end_col, chart_position):

    chart = LineChart()
    chart.title = chart_name
    chart.style = 10
    chart.height = 10
    chart.width = 20
    chart.y_axis.title = y_axis_name
    chart.x_axis.title = x_axis_name

    data = Reference(sheet, min_col = start_col, min_row=start_row, max_col = end_col, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    
    s1 = chart.series[0]
    s1.graphicalProperties.line.width = 30050
    s2 = chart.series[1]
    s2.graphicalProperties.line.width = 30050
    s3 = chart.series[2]
    s3.graphicalProperties.line.width = 30050
    s4 = chart.series[3]
    s4.graphicalProperties.line.width = 30050
    s5 = chart.series[4]
    s5.graphicalProperties.line.width = 30050
    
    sheet.add_chart(chart, chart_position)
    
for index in range(0,len(infile_names)):

    for run_name in run_names:
        print("Working on " + infile_names[index] + " and run " + run_name)
        infile = os.path.join(wd, run_name, "output", infile_names[index])
        outfile = os.path.join(wd, "compare", outfile_names[index])
        #outfile1 = os.path.join(wd, "compare", "test.xlsx")

        print('Reopen file and save again')
        xl = DispatchEx('Excel.Application')
        xl.Visible = False
        wb = xl.Workbooks.Open(infile)
        wb.Close(True)

        inworkbook = xlrd.open_workbook(infile)

        if (infile_names[index]=="TripDestination.xlsm"):
            
            outworkbook = openpyxl.load_workbook(outfile, keep_vba=True) #, keep_vba=True
            
            for purp in purpose:
                print(purp)
                sheet_name = "TLFD_" + purp
                insheet = inworkbook.sheet_by_name(sheet_name)
                nrows = insheet.nrows

                outsheet = outworkbook.get_sheet_by_name(sheet_name)

                #------- Avg Trip Length
                #column number for avg trip length - openpyxl- row and col starts at 1
                if run_name == 'soundcast':
                    col=18
                    col_all = 25
                elif run_name=='bkrcast_all':
                    col=15
                    col_all = 26
                    #add survey avg. trip length
                    value = insheet.cell_value(10,10) #xlrd- row and col starts at 0
                    outsheet.cell(row=11, column=14).value = value #survey avg trip length
                    if (purp=='Home'):
                        #add overall avg trip length in survey
                        value = insheet.cell_value(18,21) #xlrd- row and col starts at 0
                        outsheet.cell(row=13, column=24).value = value #survey avg trip length
                    
                elif run_name=='bkrcast_inbkr':
                    col=16
                    col_all = 27
                elif run_name=='bkrcast_outbkr':
                    col=17
                    col_all = 28

                value = insheet.cell_value(10,11) #xlrd- row and col starts at 0
                outsheet.cell(row=11, column=col).value = value #daysim avg trip length

                if (purp=='Home' and run_name <> 'soundcast'):
                    #add overall avg trip length
                    value = insheet.cell_value(18,25) #xlrd- row and col starts at 0
                    outsheet.cell(row=13, column=col_all).value = value #survey avg trip length                

                #------- Trip Length Frequency Distribution
                for nrow in range(0,nrows):
                    #column number for trip length frequencies
                    if run_name == 'soundcast':
                        col=10
                    elif run_name=='bkrcast_all':
                        col=7
                    elif run_name=='bkrcast_inbkr':
                        col=8
                    elif run_name=='bkrcast_outbkr':
                        col=9

                    if (nrow<>9):  #header                
                        #get data 
                        data = [insheet.cell_value(nrow,6)]
                        data_survey = [insheet.cell_value(nrow,2)]

                        #replace data
                        for column, value in enumerate(data):
                            outsheet.cell(row=nrow+1, column=col).value = value

                        if run_name=='bkrcast_all':
                            #survey data
                            for column, value in enumerate(data_survey):
                                outsheet.cell(row=nrow+1, column=3).value = value
                        
            outworkbook.save(outfile)

        else:
            insheet = inworkbook.sheet_by_name(sheet_names[index])
            nrows = insheet.nrows

            #inworkbook = openpyxl.load_workbook(infile, data_only=True,keep_vba=True)
            #insheet = inworkbook.get_sheet_by_name(sheet_names[index])
            #nrows = insheet.max_row

            outworkbook = openpyxl.load_workbook(outfile, keep_vba=True) #, keep_vba=True
            outsheet = outworkbook.get_sheet_by_name(run_name)

            for nrow in range(0,nrows):
                #get data
                data = [insheet.cell_value(nrow,col) for col in range(insheet.ncols)]
                #data = [insheet.cell(row=nrow+1,column=col+1).value for col in range(insheet.max_column)]

                #replace data
                for col, value in enumerate(data):
                    outsheet.cell(row=nrow+1, column=col+1).value = value

            outworkbook.save(outfile)
            
        if (infile_names[index]=="WrkLocation.xlsm" or infile_names[index]=="SchLocation.xlsm"):
            
            outworkbook = openpyxl.load_workbook(outfile, keep_vba=True) #, keep_vba=True
            outsheet = outworkbook.get_sheet_by_name("TLFD")

            print("adding chart")
            
            if (infile_names[index]=="WrkLocation.xlsm"):
                add_chart(outsheet, "Home to Work Distance", "Distance (mile)", "% Workers", 10, 100, 15, 19 ,"U18")
            else:
                add_chart(outsheet, "Home to School Distance", "Distance (mile)", "% Students", 10, 100, 15, 19 ,"U18")
                
            outworkbook.save(outfile)

        elif (infile_names[index]=="TourDestination_Escort.xlsm" or infile_names[index]=="TourDestination_Meal.xlsm"
              or infile_names[index]=="TourDestination_PerBus.xlsm" or infile_names[index]=="TourDestination_Shop.xlsm"
              or infile_names[index]=="TourDestination_SocRec.xlsm"):
            
            outworkbook = openpyxl.load_workbook(outfile, keep_vba=True) #, keep_vba=True
            outsheet = outworkbook.get_sheet_by_name("TLFD")

            print("adding chart")
            add_chart(outsheet, "Tour Length Frequency Distribution", "Tour Length (mile)", "% Tours", 10, 100, 7, 11 ,"M15")
                
            outworkbook.save(outfile)
            
        elif (infile_names[index]=="TourDestination_WrkBased.xlsm"):
            outworkbook = openpyxl.load_workbook(outfile, keep_vba=True) #, keep_vba=True
            outsheet = outworkbook.get_sheet_by_name("TLFD")

            print("adding chart")
            add_chart(outsheet, "Tour Length Frequency Distribution", "Tour Length (mile)", "% Tours", 10, 100, 13, 17 ,"S18")
                
            outworkbook.save(outfile)
            
        elif (infile_names[index]=="TripDestination.xlsm"):
            outworkbook = openpyxl.load_workbook(outfile, keep_vba=True) #, keep_vba=True
            for purp in purpose:
                print(purp)
                sheet_name = "TLFD_" + purp
                outsheet = outworkbook.get_sheet_by_name(sheet_name)

                print("adding chart")
                add_chart(outsheet, purp, "Trip Length (mile)", "% Trips", 10, 100, 7, 11 ,"M15")
                    
            outworkbook.save(outfile)            
